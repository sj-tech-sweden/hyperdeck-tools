from __future__ import annotations

from datetime import date, datetime, time
from typing import Any

from openpyxl import load_workbook

PLUGIN_LABEL = "Excel Schedule Upload"
PLUGIN_DESCRIPTION = "Upload an .xlsx file and convert rows into schedule entries."
PLUGIN_SUPPORTS_FILE_UPLOAD = True


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _format_datetime(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return datetime.combine(value, time.min).strftime("%Y-%m-%d %H:%M")
    return str(value or "").strip()


def _build_start_time(row: dict[str, Any]) -> str:
    direct = row.get("start_time") or row.get("start") or row.get("datetime")
    if direct:
        return _format_datetime(direct)

    date_part = row.get("date")
    time_part = row.get("time")
    if date_part and time_part:
        d = _format_datetime(date_part).split(" ")[0]
        t = str(time_part).strip()[:5]
        return f"{d} {t}"
    return ""


def scrape(file_path: str = "app/backend/uploads/schedule.xlsx") -> list[dict[str, str]]:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_norm_header(h) for h in rows[0]]
    items: list[dict[str, str]] = []

    for row_values in rows[1:]:
        raw = {headers[idx]: row_values[idx] for idx in range(min(len(headers), len(row_values)))}

        start_time = _build_start_time(raw)
        planned_title = str(
            raw.get("planned_title")
            or raw.get("title")
            or raw.get("event")
            or ""
        ).strip()
        stage = str(raw.get("stage") or raw.get("venue") or "").strip()
        row_id = str(raw.get("id") or "").strip()

        if not planned_title and not start_time:
            continue

        if not row_id:
            title_token = planned_title.replace(" ", "_").lower() or "event"
            time_token = start_time.replace(" ", "_").replace(":", "") if start_time else "unscheduled"
            row_id = f"{time_token}_{title_token}"

        item: dict[str, str] = {
            "id": row_id,
            "planned_title": planned_title or row_id,
        }
        if start_time:
            item["start_time"] = start_time
        if stage:
            item["stage"] = stage

        items.append(item)

    return items
